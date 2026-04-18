"""
Générateur de rapport PEI — Points d'Eau Incendie
Source : export TimeTonic (xlsx ou csv)
Usage  : python3 generate_rapport_pei.py <source.xlsx> <rapport.xlsx>
"""

import sys, io
import pandas as pd
import numpy as np
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from matplotlib.ticker import PercentFormatter

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

# ─── CONFIG ─────────────────────────────────────────────
# Taille des graphiques dans Excel en EMU (English Metric Units)
# 1 cm = 360000 EMU  |  1 pouce = 914400 EMU
# A4 paysage imprimable ≈ 25 cm de large
CM = 360000   # 1 cm en EMU

# Résolution matplotlib : grande figure + haut DPI = image nette
DPI = 220

C_CONF   = '#27AE60'
C_NCONF  = '#E67E22'
C_HORS   = '#E74C3C'
C_PRES_M = '#E74C3C'   # < 1 bar
C_PRES_P = '#27AE60'   # >= 1 bar
C_PRES_N = '#BDC3C7'   # non mesurée
C_ANOM   = '#E74C3C'

CONF_ORDER  = ['Conforme', 'Non conforme', 'Hors service']
CONF_COLORS = [C_CONF, C_NCONF, C_HORS]

PALETTE_XL = {
    'BLEU':       'FF1B3A5C',
    'BLEU_CLAIR': 'FFD6E4F0',
    'ACCENT':     'FF2E86AB',
    'VERT':       'FF27AE60',
    'ORANGE':     'FFE67E22',
    'ROUGE':      'FFE74C3C',
    'VIOLET':     'FF6C3483',
    'GRIS':       'FFF5F5F5',
    'BLANC':      'FFFFFFFF',
    'DARK':       'FF2C3E50',
}

plt.rcParams.update({
    'font.family':    'DejaVu Sans',
    'font.size':      11,
    'axes.titlesize': 12,
    'axes.titleweight': 'bold',
    'axes.titlecolor': '#1B3A5C',
    'figure.facecolor': 'white',
    'axes.facecolor':   'white',
    'axes.spines.top':   False,
    'axes.spines.right': False,
})

# ─── HELPERS EXCEL ──────────────────────────────────────

def xfill(hex_c):
    return PatternFill('solid', start_color=hex_c, end_color=hex_c)

def xborder(color='FFB8B8B8'):
    s = Side(style='thin', color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def xcell(ws, coord, value='', bg='FFFFFFFF', fg='FF1A1A1A', sz=10,
          bold=False, align='left', wrap=False, border=True):
    c = ws[coord]
    c.value = value
    c.font = Font(name='Arial', size=sz, bold=bold, color=fg)
    c.fill = xfill(bg)
    c.alignment = Alignment(horizontal=align, vertical='center', wrap_text=wrap)
    if border:
        c.border = xborder()
    return c

def title_band(ws, coord, value, merge_to, bg=None, sz=13):
    bg = bg or PALETTE_XL['BLEU']
    ws.merge_cells(f'{coord}:{merge_to}')
    c = ws[coord]
    c.value = value
    c.font = Font(name='Arial', size=sz, bold=True, color='FFFFFFFF')
    c.fill = xfill(bg)
    c.alignment = Alignment(horizontal='center', vertical='center')
    return c

def section_hdr(ws, coord, value, merge_to, bg=None):
    bg = bg or PALETTE_XL['ACCENT']
    ws.merge_cells(f'{coord}:{merge_to}')
    c = ws[coord]
    c.value = value
    c.font = Font(name='Arial', size=10, bold=True, color='FFFFFFFF')
    c.fill = xfill(bg)
    c.alignment = Alignment(horizontal='left', vertical='center')
    c.border = xborder()
    return c

def kpi(ws, row, col, label, value, color_hex):
    """2 lignes : label gris clair + valeur colorée."""
    c1 = ws[f'{col}{row}']
    c1.value = label
    c1.font = Font(name='Arial', size=9, color='FF555555')
    c1.fill = xfill(PALETTE_XL['GRIS'])
    c1.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c1.border = xborder()

    c2 = ws[f'{col}{row+1}']
    c2.value = value
    c2.font = Font(name='Arial', size=22, bold=True, color=color_hex.replace('#','FF'))
    c2.fill = xfill(PALETTE_XL['BLANC'])
    c2.alignment = Alignment(horizontal='center', vertical='center')
    c2.border = xborder()

def grey_block(ws, r1, r2, cols):
    for r in range(r1, r2+1):
        for col in cols:
            ws[f'{col}{r}'].fill = xfill(PALETTE_XL['GRIS'])

def insert_img(ws, buf, anchor, w_cm, h_cm):
    """Insère une image en fixant la taille en EMU (indépendant du zoom Excel)."""
    buf.seek(0)
    img = XLImage(buf)
    img.width  = int(w_cm * CM / 9144)   # conversion EMU → pixels à 96dpi
    img.height = int(h_cm * CM / 9144)
    ws.add_image(img, anchor)

def rows_height(ws, r1, r2, h_pt):
    for r in range(r1, r2+1):
        ws.row_dimensions[r].height = h_pt

# ─── GRAPHIQUES MATPLOTLIB ───────────────────────────────

def fig_to_buf(fig):
    buf = io.BytesIO()
    fig.savefig(buf, format='png', dpi=DPI, bbox_inches='tight',
                facecolor='white', edgecolor='none')
    plt.close(fig)
    buf.seek(0)
    return buf

def chart_pie(data_dict, title, colors):
    """Camembert propre avec % et légende."""
    labels = list(data_dict.keys())
    values = list(data_dict.values())
    fig, ax = plt.subplots(figsize=(5.5, 4.5))
    wedges, _, autotexts = ax.pie(
        values, colors=colors, startangle=90,
        autopct=lambda p: f'{p:.1f}%',
        pctdistance=0.72,
        wedgeprops={'edgecolor': 'white', 'linewidth': 2.5}
    )
    for at in autotexts:
        at.set_fontsize(11)
        at.set_fontweight('bold')
        at.set_color('white')
    ax.legend(wedges, [f'{l} ({v})' for l, v in zip(labels, values)],
              loc='lower center', bbox_to_anchor=(0.5, -0.12),
              ncol=len(labels), fontsize=9, frameon=False)
    ax.set_title(title, pad=12)
    fig.tight_layout()
    return fig_to_buf(fig)

def chart_stacked(df, col, title):
    """Barres horizontales empilées 100% avec % dans chaque segment."""
    cats = sorted(df[col].dropna().unique())
    ct = pd.crosstab(df[col], df['Conformité'])
    for s in CONF_ORDER:
        if s not in ct.columns:
            ct[s] = 0
    ct = ct[CONF_ORDER]
    ct_pct = ct.div(ct.sum(axis=1), axis=0) * 100

    n = len(cats)
    fig, ax = plt.subplots(figsize=(7, max(2.2, n * 1.1)))
    lefts = np.zeros(n)

    for statut, color in zip(CONF_ORDER, CONF_COLORS):
        vals   = ct_pct.loc[cats, statut].values
        counts = ct.loc[cats, statut].values
        bars = ax.barh(cats, vals, left=lefts, color=color,
                       height=0.55, edgecolor='white', linewidth=1.5,
                       label=statut)
        for bar, pct, cnt in zip(bars, vals, counts):
            if pct > 9:
                ax.text(
                    bar.get_x() + bar.get_width() / 2,
                    bar.get_y() + bar.get_height() / 2,
                    f'{pct:.0f}%\n({cnt})',
                    ha='center', va='center',
                    fontsize=9, fontweight='bold', color='white'
                )
        lefts += vals

    ax.set_xlim(0, 100)
    ax.set_title(title, pad=10)
    ax.tick_params(axis='y', labelsize=10)
    ax.set_xticks([])
    ax.set_xlabel('')
    ax.spines['left'].set_visible(False)
    ax.spines['bottom'].set_visible(False)
    ax.invert_yaxis()

    patches = [mpatches.Patch(color=c, label=s) for s, c in zip(CONF_ORDER, CONF_COLORS)]
    ax.legend(handles=patches, loc='lower center', bbox_to_anchor=(0.5, -0.22),
              ncol=3, fontsize=9, frameon=False)

    fig.subplots_adjust(bottom=0.18)
    fig.tight_layout(rect=[0, 0.12, 1, 1])
    return fig_to_buf(fig)

def chart_obs(obs_counts, title):
    """Barres horizontales top anomalies, propres, sans gridlines."""
    exploded = {}
    for obs_str, count in obs_counts.items():
        if isinstance(obs_str, str) and obs_str.strip().upper() != 'RAS':
            for item in obs_str.split(','):
                item = item.strip()
                if item:
                    exploded[item] = exploded.get(item, 0) + count
    top = sorted(exploded.items(), key=lambda x: x[1], reverse=True)[:10]
    if not top:
        return None

    labels = [t[0] for t in top]
    values = [t[1] for t in top]

    fig, ax = plt.subplots(figsize=(9, max(3.5, len(labels) * 0.7)))
    bars = ax.barh(labels, values, color=C_ANOM, height=0.55,
                   edgecolor='white', linewidth=1)
    for bar, val in zip(bars, values):
        ax.text(bar.get_width() + 0.4,
                bar.get_y() + bar.get_height() / 2,
                str(val), va='center', fontsize=10,
                fontweight='bold', color='#333333')
    ax.set_xlim(0, max(values) * 1.18)
    ax.set_xlabel('Occurrences', fontsize=9)
    ax.set_title(title, pad=10)
    ax.tick_params(axis='y', labelsize=9)
    ax.set_xticks([])
    ax.spines['left'].set_visible(False)
    ax.spines['bottom'].set_visible(False)
    ax.invert_yaxis()
    fig.tight_layout()
    return fig_to_buf(fig)

# ─── MAIN ────────────────────────────────────────────────

def build_rapport_pei(source_path, output_path):
    # Chargement
    df = pd.read_csv(source_path) if source_path.endswith('.csv') \
         else pd.read_excel(source_path)

    col_map = {
        'Saisir le N° du PEI':       'N° PEI',
        'Choisissez la commune':      'Commune',
        "Localisation de l'Hydrant":  'Localisation',
        'Fabricant':                  'Fabricant',
        'Famille':                    'Famille',
        'Type':                       'Type',
        'Version':                    'Version',
        'Technicien':                 'Technicien',
        'Date':                       'Date',
        'Heure':                      'Heure',
        'Pression statique (en bar)': 'Pression',
        'Débit à 1 bar (m3/h)':      'Débit',
        'Conformité':                 'Conformité',
        "Etat de l'hydrant":          'Etat',
        'Observation':                'Observation',
        'Autres observations':        'Autres obs',
    }
    df = df.rename(columns={k: v for k, v in col_map.items() if k in df.columns})

    commune    = df['Commune'].dropna().iloc[0]    if 'Commune'    in df.columns else 'Commune'
    technicien = df['Technicien'].dropna().iloc[0] if 'Technicien' in df.columns else ''
    total      = len(df)

    conf      = df['Conformité'].value_counts().to_dict()
    conforme  = conf.get('Conforme', 0)
    non_conf  = conf.get('Non conforme', 0)
    hors_serv = conf.get('Hors service', 0)

    pression  = df['Pression'].dropna() if 'Pression' in df.columns else pd.Series([], dtype=float)
    p_moins1  = int((pression < 1).sum())
    p_plus1   = int((pression >= 1).sum())
    p_no_mes  = int(df['Pression'].isna().sum()) if 'Pression' in df.columns else 0

    anomalies   = df[df['Etat'].str.contains('Avec anomalie', na=False)] \
                  if 'Etat' in df.columns else pd.DataFrame()
    debit_insuf = df[df['Observation'].str.contains('Débit insuf', na=False)] \
                  if 'Observation' in df.columns else pd.DataFrame()
    p_anomalie  = int(len(anomalies[anomalies['Pression'].fillna(99) < 1])) \
                  if 'Pression' in anomalies.columns else 0
    obs_counts  = df['Observation'].value_counts().to_dict() \
                  if 'Observation' in df.columns else {}

    # ── Génération des graphiques ──
    print('  Génération des graphiques...')
    buf_conf  = chart_pie(
        {'Conforme': conforme, 'Non conforme': non_conf, 'Hors service': hors_serv},
        'Conformité des PEI', CONF_COLORS
    )
    buf_pres  = chart_pie(
        {'< 1 bar': p_moins1, '≥ 1 bar': p_plus1, 'Non mesurée': p_no_mes},
        'Pression statique', [C_PRES_M, C_PRES_P, C_PRES_N]
    )
    buf_fam   = chart_stacked(df, 'Famille', 'Famille × Conformité') \
                if 'Famille'  in df.columns else None
    buf_type  = chart_stacked(df, 'Type',    'Type × Conformité') \
                if 'Type'     in df.columns else None
    buf_ver   = chart_stacked(df, 'Version', 'Version × Conformité') \
                if 'Version'  in df.columns else None
    buf_obs   = chart_obs(obs_counts, 'Top anomalies observées')

    # ── Workbook ──
    wb = Workbook()

    # ════════════════════════════════════════
    # ONGLET SYNTHÈSE
    # ════════════════════════════════════════
    ws = wb.active
    ws.title = 'Synthèse'
    ws.sheet_view.showGridLines = False

    # Mise en page impression A4 paysage
    from openpyxl.worksheet.page import PageMargins
    ws.page_setup.orientation  = 'landscape'
    ws.page_setup.paperSize    = 9
    ws.page_setup.fitToPage    = True
    ws.page_setup.fitToWidth   = 1
    ws.page_setup.fitToHeight  = 0
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.6, bottom=0.6,
                                  header=0.2, footer=0.2)

    # Colonnes : A marge | B large | C-G égales | H marge
    for col, w in [('A',1),('B',24),('C',17),('D',17),('E',17),('F',17),('G',17),('H',1)]:
        ws.column_dimensions[col].width = w

    # ── PAGE 1 ──────────────────────────────

    # Titre
    ws.row_dimensions[1].height = 42
    title_band(ws, 'B1', f"RAPPORT DE CONTRÔLE DES POINTS D'EAU INCENDIE — {commune.upper()}",
               'G1', sz=13)

    ws.row_dimensions[2].height = 18
    ws.merge_cells('B2:G2')
    c = ws['B2']
    c.value = (f"Technicien : {technicien}   |   "
               f"Données au {pd.Timestamp.now().strftime('%d/%m/%Y')}   |   "
               f"Source : export TimeTonic")
    c.font  = Font(name='Arial', size=9, italic=True, color='FF555555')
    c.fill  = xfill(PALETTE_XL['BLEU_CLAIR'])
    c.alignment = Alignment(horizontal='center', vertical='center')

    # Bloc 1 — Conformité
    ws.row_dimensions[3].height = 6
    ws.row_dimensions[4].height = 20
    section_hdr(ws, 'B4', '  CONFORMITÉ DES PEI', 'G4', PALETTE_XL['BLEU'])
    ws.row_dimensions[5].height = 22
    ws.row_dimensions[6].height = 36
    for col_l, lbl, val, col_hex in [
        ('B', 'Total PEI\ncontrôlés', total,    '#1B3A5C'),
        ('C', 'Conformes',            conforme,  '#27AE60'),
        ('D', 'Non conformes',        non_conf,  '#E67E22'),
        ('E', 'Hors service',         hors_serv, '#E74C3C'),
    ]:
        kpi(ws, 5, col_l, lbl, val, col_hex)
    grey_block(ws, 5, 6, ['F', 'G'])

    # Bloc 2 — Anomalies
    ws.row_dimensions[7].height = 6
    ws.row_dimensions[8].height = 20
    section_hdr(ws, 'B8', '  ANOMALIES', 'G8', PALETTE_XL['VIOLET'])
    ws.row_dimensions[9].height = 22
    ws.row_dimensions[10].height = 36
    for col_l, lbl, val, col_hex in [
        ('B', 'PEI avec\nanomalies',     len(anomalies),  '#8E44AD'),
        ('C', 'dont Débit\ninsuffisant', len(debit_insuf),'#E74C3C'),
        ('D', 'dont Pression\n< 1 bar',  p_anomalie,      '#E67E22'),
    ]:
        kpi(ws, 9, col_l, lbl, val, col_hex)
    grey_block(ws, 9, 10, ['E', 'F', 'G'])

    # Section répartition P1
    ws.row_dimensions[11].height = 6
    ws.row_dimensions[12].height = 20
    section_hdr(ws, 'B12', '  RÉPARTITION DES HYDRANTS', 'G12')

    # Graphiques ligne 1 : camembert (11cm) + famille (14cm) — hauteur 9.5cm
    rows_height(ws, 13, 27, 20)
    insert_img(ws, buf_conf, 'B13', w_cm=11.5, h_cm=9.5)
    if buf_fam:
        insert_img(ws, buf_fam, 'E13', w_cm=14.5, h_cm=9.5)

    # Section type & version
    ws.row_dimensions[28].height = 6
    ws.row_dimensions[29].height = 20
    section_hdr(ws, 'B29', '  TYPE & VERSION', 'G29')

    # Graphiques ligne 2 : type (12.5cm) + version (12.5cm) — hauteur 8cm
    rows_height(ws, 30, 41, 20)
    if buf_type:
        insert_img(ws, buf_type, 'B30', w_cm=12.5, h_cm=8.0)
    if buf_ver:
        insert_img(ws, buf_ver,  'E30', w_cm=12.5, h_cm=8.0)

    # Saut de page après ligne 41
    from openpyxl.worksheet.pagebreak import Break
    ws.row_breaks.append(Break(id=41))

    # ── PAGE 2 ──────────────────────────────

    ws.row_dimensions[42].height = 20
    section_hdr(ws, 'B42', '  PRESSION STATIQUE', 'G42')

    rows_height(ws, 43, 55, 20)
    insert_img(ws, buf_pres, 'B43', w_cm=11.5, h_cm=9.5)

    ws.row_dimensions[56].height = 6
    ws.row_dimensions[57].height = 20
    section_hdr(ws, 'B57', '  TOP ANOMALIES OBSERVÉES', 'G57')

    rows_height(ws, 58, 80, 20)
    if buf_obs:
        insert_img(ws, buf_obs, 'B58', w_cm=25.0, h_cm=14.0)

    # ════════════════════════════════════════
    # ONGLETS DÉTAIL PAR CONFORMITÉ
    # ════════════════════════════════════════
    for statut, bg_hex in [
        ('Conforme',     PALETTE_XL['VERT']),
        ('Non conforme', PALETTE_XL['ORANGE']),
        ('Hors service', PALETTE_XL['ROUGE']),
    ]:
        wsc = wb.create_sheet(title=statut)
        wsc.sheet_view.showGridLines = False
        for c_l, w in [('A',1),('B',8),('C',22),('D',40),('E',12),
                       ('F',12),('G',12),('H',12),('I',18),('J',30),('K',1)]:
            wsc.column_dimensions[c_l].width = w

        wsc.row_dimensions[2].height = 38
        title_band(wsc, 'B2', f'{statut.upper()} — {commune}', 'J2', bg_hex, sz=13)

        wsc.row_dimensions[3].height = 6
        wsc.row_dimensions[4].height = 20
        hdrs = ['N° PEI','Localisation','Famille','Fabricant','Type',
                'Pression (bar)','Débit (m3/h)','Etat','Observation']
        c_ls = ['B','C','D','E','F','G','H','I','J']
        for c_l, hdr in zip(c_ls, hdrs):
            xcell(wsc, f'{c_l}4', hdr, bg=PALETTE_XL['DARK'],
                  fg='FFFFFFFF', sz=9, bold=True, align='center')

        subset = df[df['Conformité'] == statut].reset_index(drop=True)
        for i, row_data in subset.iterrows():
            r = 5 + i
            wsc.row_dimensions[r].height = 18
            bg = 'FFF7F7F7' if i % 2 == 0 else PALETTE_XL['BLANC']
            vals = [row_data.get(k, '') for k in
                    ['N° PEI','Localisation','Famille','Fabricant','Type',
                     'Pression','Débit','Etat','Observation']]
            for c_l, val in zip(c_ls, vals):
                align = 'center' if c_l in ['B','E','F','G','H'] else 'left'
                xcell(wsc, f'{c_l}{r}', val if pd.notna(val) else '',
                      bg=bg, sz=9, align=align,
                      wrap=(c_l in ['D','I','J']))

        last = 5 + len(subset)
        wsc.merge_cells(f'B{last}:J{last}')
        c = wsc[f'B{last}']
        c.value = f'Total : {len(subset)} PEI {statut.lower()}(s)'
        c.font  = Font(name='Arial', size=10, bold=True, color='FFFFFFFF')
        c.fill  = xfill(bg_hex)
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = xborder()

    # ════════════════════════════════════════
    # ONGLET DONNÉES COMPLÈTES
    # ════════════════════════════════════════
    wsd = wb.create_sheet(title='Données complètes')
    wsd.sheet_view.showGridLines = False

    exp_cols = ['N° PEI','Localisation','Famille','Fabricant','Type','Version',
                'Date','Technicien','Pression','Débit','Conformité','Etat',
                'Observation','Autres obs']
    col_lts  = [get_column_letter(i+2) for i in range(len(exp_cols))]

    wsd.row_dimensions[2].height = 38
    title_band(wsd, 'B2', f'DONNÉES COMPLÈTES — {commune.upper()}',
               col_lts[-1]+'2', sz=13)

    wsd.row_dimensions[3].height = 6
    wsd.row_dimensions[4].height = 20
    for c_l, hdr in zip(col_lts, exp_cols):
        xcell(wsd, f'{c_l}4', hdr, bg=PALETTE_XL['DARK'],
              fg='FFFFFFFF', sz=9, bold=True, align='center')
        wsd.column_dimensions[c_l].width = max(10, len(hdr)+4)
    wsd.column_dimensions['C'].width = 42
    wsd.column_dimensions[col_lts[12]].width = 32

    for i, row_data in df.iterrows():
        r = 5 + i
        wsd.row_dimensions[r].height = 16
        bg = 'FFF7F7F7' if i % 2 == 0 else PALETTE_XL['BLANC']
        for c_l, col_key in zip(col_lts, exp_cols):
            val     = row_data.get(col_key, '')
            val_str = val if pd.notna(val) else ''
            bg_c, fg_c = bg, 'FF1A1A1A'
            if col_key == 'Conformité':
                if val_str == 'Conforme':
                    bg_c, fg_c = 'FFD5F5E3', 'FF1E8449'
                elif val_str == 'Non conforme':
                    bg_c, fg_c = 'FFFDE8D8', 'FFA04000'
                elif val_str == 'Hors service':
                    bg_c, fg_c = 'FFFDEDEC', 'FF922B21'
            align = 'center' if c_l in [col_lts[0], col_lts[4], col_lts[5]] else 'left'
            xcell(wsd, f'{c_l}{r}', val_str, bg=bg_c, fg=fg_c,
                  sz=9, align=align,
                  wrap=(col_key in ['Observation','Autres obs','Localisation']))

    wb.save(output_path)
    print(f"✅ Rapport généré : {output_path}")
    print(f"   {commune} | {total} PEI | "
          f"Conforme: {conforme} | Non conforme: {non_conf} | Hors service: {hors_serv}")

# ─────────────────────────────────────────────
if __name__ == '__main__':
    src = sys.argv[1] if len(sys.argv) > 1 else None
    out = sys.argv[2] if len(sys.argv) > 2 else 'rapport_pei.xlsx'
    if not src:
        print("Usage: python3 generate_rapport_pei.py <source.xlsx|csv> <rapport.xlsx>")
        sys.exit(1)
    build_rapport_pei(src, out)
