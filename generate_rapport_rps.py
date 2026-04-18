import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.chart import BarChart, RadarChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import numpy as np
import io, sys, os

# ─────────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────────
REPONSES_ORDER = ['Jamais, non', 'Parfois, plutôt non', 'Souvent, plutôt oui', 'Toujours oui']
SCORE_MAP = {'Jamais, non': 1, 'Parfois, plutôt non': 2, 'Souvent, plutôt oui': 3, 'Toujours oui': 4}

COULEURS_REPONSES = {
    'Jamais, non':        '#2ecc71',
    'Parfois, plutôt non':'#f1c40f',
    'Souvent, plutôt oui':'#e67e22',
    'Toujours oui':       '#e74c3c',
}

THEMES = {
    'Charge & intensité': [1, 6, 7, 8, 17],
    'Organisation & clarté': [2, 3, 4, 5, 18],
    'Autonomie & marges': [15, 27, 28],
    'Équilibre vie pro/perso': [9, 10, 11],
    'Relations & soutien': [24, 25, 26],
    'Reconnaissance & sens': [20, 22, 23],
    'Tensions émotionnelles': [12, 13, 14],
    'Bien-être & stress': [16, 19, 21],
}

QUESTIONS_TEXTES = {
    1: "Contraintes de rythmes élevés ?",
    2: "Objectifs clairement définis ?",
    3: "Objectifs compatibles avec les moyens ?",
    4: "Instructions contradictoires ?",
    5: "Changements de tâches à l'improviste ?",
    6: "Interruptions par tâches non prévues ?",
    7: "Activités à vigilance permanente ?",
    8: "Plus de 45h/semaine ?",
    9: "Horaires de nuit/alternants/décalés ?",
    10: "Contacté hors horaires de travail ?",
    11: "Conciliation vie pro/perso possible ?",
    12: "Organisation génère tensions avec public ?",
    13: "Moyens face à la souffrance d'autrui ?",
    14: "Faire bonne figure en toutes circonstances ?",
    15: "Marges de manœuvre dans la méthode ?",
    16: "Épanouissement au travail ?",
    17: "Charge de travail lourde ?",
    18: "Changements anticipés et expliqués ?",
    19: "Stress au travail ?",
    20: "Travail reconnu comme utile ?",
    21: "Incertitudes sur le maintien de l'activité ?",
    22: "Marques de reconnaissance de l'entreprise ?",
    23: "Travail de qualité réalisé ?",
    24: "Désaccords liés à l'organisation ?",
    25: "Possibilités d'entraide entre salariés ?",
    26: "Soutien de l'encadrement ?",
    27: "Possibilité d'interrompre son travail ?",
    28: "Utilisation et développement des compétences ?",
}

# Palette couleurs rapport
C_BLEU_FOND   = 'FF1B3A5C'
C_BLEU_CLAIR  = 'FFD6E4F0'
C_ACCENT      = 'FF2E86AB'
C_VERT        = 'FF2ECC71'
C_JAUNE       = 'FFF1C40F'
C_ORANGE      = 'FFE67E22'
C_ROUGE       = 'FFE74C3C'
C_GRIS        = 'FFF5F5F5'
C_BLANC       = 'FFFFFFFF'

def side(style='thin', color='FFB0B0B0'):
    return Side(style=style, color=color)

def border_full(style='thin'):
    s = side(style)
    return Border(left=s, right=s, top=s, bottom=s)

def fill(hex_color):
    return PatternFill('solid', start_color=hex_color, end_color=hex_color)

def normalize_response(r):
    if not isinstance(r, str):
        return None
    r = r.strip().lower()
    for ref in REPONSES_ORDER:
        if ref.lower() == r:
            return ref
    # fuzzy
    if 'jamais' in r: return 'Jamais, non'
    if 'parfois' in r: return 'Parfois, plutôt non'
    if 'souvent' in r: return 'Souvent, plutôt oui'
    if 'toujours' in r: return 'Toujours oui'
    return None

def score_risque(score_norm):
    """score_norm: 1-4, retourne label et couleur hex"""
    if score_norm <= 1.75: return 'Faible', '#2ecc71'
    if score_norm <= 2.5:  return 'Modéré', '#f1c40f'
    if score_norm <= 3.25: return 'Élevé', '#e67e22'
    return 'Critique', '#e74c3c'

# ─────────────────────────────────────────────
# MATPLOTLIB CHARTS → PNG in memory
# ─────────────────────────────────────────────

def make_stacked_bar(df, questions, theme_name):
    """Horizontal stacked bar chart for a theme."""
    labels = [f"Q{q} – {QUESTIONS_TEXTES[q]}" for q in questions]
    colors = [COULEURS_REPONSES[r] for r in REPONSES_ORDER]

    counts = []
    for q in questions:
        col = f'Q{q}'
        total = len(df)
        row = []
        for r in REPONSES_ORDER:
            n = (df[col] == r).sum()
            row.append(round(n / total * 100, 1))
        counts.append(row)

    counts = np.array(counts)
    fig, ax = plt.subplots(figsize=(10, max(2.5, len(questions) * 0.85)))
    fig.patch.set_facecolor('#FAFAFA')
    ax.set_facecolor('#FAFAFA')

    lefts = np.zeros(len(questions))
    for i, (rep, color) in enumerate(zip(REPONSES_ORDER, colors)):
        bars = ax.barh(labels, counts[:, i], left=lefts, color=color,
                       label=rep, height=0.55, edgecolor='white', linewidth=0.5)
        # add % labels if > 8%
        for bar, val in zip(bars, counts[:, i]):
            if val > 8:
                ax.text(bar.get_x() + bar.get_width() / 2,
                        bar.get_y() + bar.get_height() / 2,
                        f'{val:.0f}%', ha='center', va='center',
                        fontsize=8, color='white', fontweight='bold')
        lefts += counts[:, i]

    ax.set_xlim(0, 100)
    ax.set_xlabel('% des répondants', fontsize=9)
    ax.set_title(theme_name, fontsize=11, fontweight='bold', pad=10, color='#1B3A5C')
    ax.legend(loc='lower right', fontsize=8, framealpha=0.8)
    ax.tick_params(axis='y', labelsize=8.5)
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)
    ax.invert_yaxis()
    plt.tight_layout()

    buf = io.BytesIO()
    plt.savefig(buf, format='png', dpi=130, bbox_inches='tight')
    plt.close()
    buf.seek(0)
    return buf

def make_radar(theme_scores):
    """Radar chart for all 8 themes."""
    themes = list(theme_scores.keys())
    scores = [theme_scores[t] for t in themes]
    N = len(themes)

    angles = np.linspace(0, 2 * np.pi, N, endpoint=False).tolist()
    scores_plot = scores + [scores[0]]
    angles += angles[:1]

    fig, ax = plt.subplots(figsize=(6, 6), subplot_kw=dict(polar=True))
    fig.patch.set_facecolor('#FAFAFA')
    ax.set_facecolor('#FAFAFA')

    # zones de couleur
    for r, c, a in [(4, '#e74c3c', 0.08), (3.25, '#e67e22', 0.10),
                    (2.5, '#f1c40f', 0.10), (1.75, '#2ecc71', 0.12)]:
        ax.fill(angles, [r] * len(angles), color=c, alpha=a)

    ax.plot(angles, scores_plot, 'o-', linewidth=2, color='#1B3A5C')
    ax.fill(angles, scores_plot, alpha=0.25, color='#2E86AB')

    ax.set_xticks(angles[:-1])
    ax.set_xticklabels(themes, size=8.5, fontweight='bold', color='#1B3A5C')
    ax.set_ylim(1, 4)
    ax.set_yticks([1, 2, 3, 4])
    ax.set_yticklabels(['1\nFaible', '2\nModéré', '3\nÉlevé', '4\nCritique'],
                       size=7, color='grey')
    ax.set_title('Vue globale par thème', fontsize=12, fontweight='bold',
                 pad=20, color='#1B3A5C')
    plt.tight_layout()

    buf = io.BytesIO()
    plt.savefig(buf, format='png', dpi=130, bbox_inches='tight')
    plt.close()
    buf.seek(0)
    return buf

def make_score_bar(theme_scores):
    """Horizontal bar showing score per theme with color coding."""
    themes = list(theme_scores.keys())
    scores = [theme_scores[t] for t in themes]
    colors = [score_risque(s)[1] for s in scores]

    fig, ax = plt.subplots(figsize=(8, 4))
    fig.patch.set_facecolor('#FAFAFA')
    ax.set_facecolor('#FAFAFA')

    bars = ax.barh(themes, scores, color=colors, height=0.55,
                   edgecolor='white', linewidth=0.5)
    for bar, score in zip(bars, scores):
        label, _ = score_risque(score)
        ax.text(bar.get_width() + 0.05, bar.get_y() + bar.get_height() / 2,
                f'{score:.2f}  [{label}]', va='center', fontsize=9,
                color='#333333')

    ax.set_xlim(1, 4.8)
    ax.set_xlabel('Score moyen (1 = Jamais → 4 = Toujours)', fontsize=9)
    ax.set_title('Score de risque par thème', fontsize=11, fontweight='bold',
                 color='#1B3A5C')
    ax.axvline(2.5, color='#e67e22', linestyle='--', linewidth=1, alpha=0.6, label='Seuil vigilance')
    ax.legend(fontsize=8)
    ax.invert_yaxis()
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)
    plt.tight_layout()

    buf = io.BytesIO()
    plt.savefig(buf, format='png', dpi=130, bbox_inches='tight')
    plt.close()
    buf.seek(0)
    return buf

# ─────────────────────────────────────────────
# EXCEL BUILDER
# ─────────────────────────────────────────────

def insert_image(ws, buf, anchor, width_px=None, height_px=None):
    img = XLImage(buf)
    if width_px: img.width = width_px
    if height_px: img.height = height_px
    ws.add_image(img, anchor)

def set_col_widths(ws, widths_dict):
    for col, w in widths_dict.items():
        ws.column_dimensions[col].width = w

def header_cell(ws, coord, text, bg=C_BLEU_FOND, fg='FFFFFFFF', sz=12,
                bold=True, align='center', wrap=False):
    c = ws[coord]
    c.value = text
    c.font = Font(name='Arial', bold=bold, size=sz, color=fg)
    c.fill = fill(bg)
    c.alignment = Alignment(horizontal=align, vertical='center', wrap_text=wrap)
    return c

def data_cell(ws, coord, text, bg=C_BLANC, fg='FF333333', sz=10,
              bold=False, align='left', wrap=False):
    c = ws[coord]
    c.value = text
    c.font = Font(name='Arial', bold=bold, size=sz, color=fg)
    c.fill = fill(bg)
    c.alignment = Alignment(horizontal=align, vertical='center', wrap_text=wrap)
    c.border = border_full()
    return c

def build_excel(csv_path, output_path, entreprise="Entreprise", n_salaries=None):
    # Load data
    df = pd.read_csv(csv_path)

    # Normalize columns
    col_map = {}
    for col in df.columns:
        stripped = col.strip()
        if stripped.upper().startswith('Q'):
            try:
                num = int(stripped[1:])
                col_map[col] = f'Q{num}'
            except:
                pass
    df = df.rename(columns=col_map)

    for q in range(1, 29):
        col = f'Q{q}'
        if col in df.columns:
            df[col] = df[col].apply(normalize_response)

    n_rep = len(df)
    if n_salaries is None:
        n_salaries = n_rep

    # Compute theme scores
    theme_scores = {}
    for theme, questions in THEMES.items():
        scores = []
        for q in questions:
            col = f'Q{q}'
            if col in df.columns:
                vals = df[col].dropna().map(SCORE_MAP)
                scores.extend(vals.tolist())
        theme_scores[theme] = round(np.mean(scores), 2) if scores else 0

    # ── Generate charts ──
    radar_buf = make_radar(theme_scores)
    score_bar_buf = make_score_bar(theme_scores)
    theme_chart_bufs = {}
    for theme, questions in THEMES.items():
        theme_chart_bufs[theme] = make_stacked_bar(df, questions, theme)

    # ── Workbook ──
    wb = Workbook()

    # ════════════════════════════════════════
    # SHEET 1 : SYNTHÈSE
    # ════════════════════════════════════════
    ws = wb.active
    ws.title = "Synthèse"
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 15
    ws.row_dimensions[2].height = 50
    ws.row_dimensions[3].height = 15

    set_col_widths(ws, {'A': 3, 'B': 28, 'C': 16, 'D': 16, 'E': 16, 'F': 16, 'G': 5})

    # Title band
    ws.merge_cells('B2:F2')
    c = ws['B2']
    c.value = f"RAPPORT DE DIAGNOSTIC RPS — {entreprise.upper()}"
    c.font = Font(name='Arial', bold=True, size=16, color='FFFFFFFF')
    c.fill = fill(C_BLEU_FOND)
    c.alignment = Alignment(horizontal='center', vertical='center')

    # Info row
    ws.row_dimensions[4].height = 22
    ws.merge_cells('B4:F4')
    c = ws['B4']
    c.value = f"Nombre de répondants : {n_rep}   |   Questionnaire anonyme   |   {pd.Timestamp.now().strftime('%B %Y')}"
    c.font = Font(name='Arial', size=10, italic=True, color='FF555555')
    c.fill = fill(C_BLEU_CLAIR)
    c.alignment = Alignment(horizontal='center', vertical='center')

    ws.row_dimensions[5].height = 10

    # Section title
    ws.row_dimensions[6].height = 22
    ws.merge_cells('B6:F6')
    header_cell(ws, 'B6', 'SCORES DE RISQUE PAR THÈME', bg=C_ACCENT, sz=11)

    # Table header
    ws.row_dimensions[7].height = 20
    for coord, txt in [('B7','Thème'),('C7','Score moyen'),('D7','Niveau de risque'),
                       ('E7','Nb questions'),('F7','Tendance')]:
        header_cell(ws, coord, txt, bg='FF2C3E50', sz=10)

    # Table rows
    for i, (theme, score) in enumerate(theme_scores.items()):
        row = 8 + i
        ws.row_dimensions[row].height = 20
        label, hex_c = score_risque(score)
        bg = 'FFF0F0F0' if i % 2 == 0 else C_BLANC
        data_cell(ws, f'B{row}', theme, bg=bg, bold=True, sz=10)
        c = ws[f'C{row}']
        c.value = score
        c.font = Font(name='Arial', bold=True, size=11, color=hex_c.replace('#','FF'))
        c.fill = fill(bg)
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = border_full()

        c2 = ws[f'D{row}']
        c2.value = label
        c2.font = Font(name='Arial', bold=True, size=10, color='FFFFFFFF')
        c2.fill = fill(hex_c.replace('#','FF'))
        c2.alignment = Alignment(horizontal='center', vertical='center')
        c2.border = border_full()

        data_cell(ws, f'E{row}', len(THEMES[theme]), bg=bg, align='center', sz=10)
        # Tendance arrow
        if score >= 3.25: arrow = '▲▲ Critique'
        elif score >= 2.5: arrow = '▲ Élevé'
        elif score >= 1.75: arrow = '► Modéré'
        else: arrow = '▼ Faible'
        data_cell(ws, f'F{row}', arrow, bg=bg, align='center', sz=10)

    # Légende
    legend_row = 8 + len(THEMES) + 1
    ws.row_dimensions[legend_row] = ws.row_dimensions[legend_row]
    ws.merge_cells(f'B{legend_row}:F{legend_row}')
    c = ws[f'B{legend_row}']
    c.value = "Légende :  ▼ Faible (1.0–1.75)   ► Modéré (1.76–2.5)   ▲ Élevé (2.51–3.25)   ▲▲ Critique (3.26–4.0)"
    c.font = Font(name='Arial', italic=True, size=9, color='FF555555')
    c.fill = fill(C_GRIS)
    c.alignment = Alignment(horizontal='center', vertical='center')

    # Charts on synthèse
    radar_row = legend_row + 2
    ws.row_dimensions[radar_row].height = 15
    insert_image(ws, radar_buf, f'B{radar_row}', width_px=380, height_px=360)
    insert_image(ws, score_bar_buf, f'E{radar_row}', width_px=480, height_px=260)

    # ════════════════════════════════════════
    # SHEET 2–9 : UN PAR THÈME
    # ════════════════════════════════════════
    for theme, questions in THEMES.items():
        safe_title = theme[:28].replace('/', '-').replace('\\', '-').replace('*', '').replace('?', '').replace('[', '').replace(']', '').replace(':', '')
        wst = wb.create_sheet(title=safe_title)
        wst.sheet_view.showGridLines = False
        set_col_widths(wst, {'A': 3, 'B': 45, 'C': 14, 'D': 14, 'E': 14, 'F': 14, 'G': 14, 'H': 5})

        # Title
        wst.row_dimensions[2].height = 40
        wst.merge_cells('B2:G2')
        c = wst['B2']
        c.value = f"THÈME : {theme.upper()}"
        c.font = Font(name='Arial', bold=True, size=14, color='FFFFFFFF')
        c.fill = fill(C_BLEU_FOND)
        c.alignment = Alignment(horizontal='center', vertical='center')

        # Score global du thème
        wst.row_dimensions[3].height = 8
        wst.row_dimensions[4].height = 22
        wst.merge_cells('B4:G4')
        score_t = theme_scores[theme]
        label_t, hex_t = score_risque(score_t)
        c = wst['B4']
        c.value = f"Score global du thème : {score_t:.2f} / 4.00   →   Niveau {label_t}"
        c.font = Font(name='Arial', bold=True, size=11, color='FFFFFFFF')
        c.fill = fill(hex_t.replace('#','FF'))
        c.alignment = Alignment(horizontal='center', vertical='center')

        wst.row_dimensions[5].height = 10

        # Table header
        wst.row_dimensions[6].height = 20
        for coord, txt in [('B6','Question'),('C6','Jamais, non'),
                           ('D6','Parfois, plutôt non'),('E6','Souvent, plutôt oui'),
                           ('F6','Toujours oui'),('G6','Score moy.')]:
            header_cell(wst, coord, txt, bg='FF2C3E50', sz=9, wrap=True)

        # Per question
        for i, q in enumerate(questions):
            row = 7 + i
            wst.row_dimensions[row].height = 22
            col = f'Q{q}'
            bg = 'FFF7F7F7' if i % 2 == 0 else C_BLANC
            data_cell(wst, f'B{row}',
                      f"Q{q} – {QUESTIONS_TEXTES[q]}", bg=bg, sz=9, wrap=True)
            total = len(df)
            scores_q = []
            for j, rep in enumerate(REPONSES_ORDER):
                n_r = (df[col] == rep).sum() if col in df.columns else 0
                pct = round(n_r / total * 100, 1)
                coord_q = f"{get_column_letter(3+j)}{row}"
                c = wst[coord_q]
                c.value = f"{n_r} ({pct}%)"
                c.font = Font(name='Arial', size=9)
                c.fill = fill(bg)
                c.alignment = Alignment(horizontal='center', vertical='center')
                c.border = border_full()
                scores_q.extend([SCORE_MAP[rep]] * n_r)
            avg_q = round(np.mean(scores_q), 2) if scores_q else 0
            label_q, hex_q = score_risque(avg_q)
            cg = wst[f'G{row}']
            cg.value = avg_q
            cg.font = Font(name='Arial', bold=True, size=10, color=hex_q.replace('#','FF'))
            cg.fill = fill(bg)
            cg.alignment = Alignment(horizontal='center', vertical='center')
            cg.border = border_full()

        # Chart
        chart_row = 7 + len(questions) + 2
        buf = theme_chart_bufs[theme]
        buf.seek(0)
        insert_image(wst, buf, f'B{chart_row}', width_px=680, height_px=max(200, len(questions)*85))

    # ════════════════════════════════════════
    # SHEET DONNÉES BRUTES
    # ════════════════════════════════════════
    wsd = wb.create_sheet(title="Données brutes")
    wsd.sheet_view.showGridLines = False
    set_col_widths(wsd, {'A': 4})
    for q in range(1, 29):
        wsd.column_dimensions[get_column_letter(q+1)].width = 22

    wsd.row_dimensions[1].height = 12
    wsd.row_dimensions[2].height = 30
    wsd.merge_cells(f'B2:{get_column_letter(29)}2')
    c = wsd['B2']
    c.value = "DONNÉES BRUTES — RÉPONSES INDIVIDUELLES (ANONYMISÉES)"
    c.font = Font(name='Arial', bold=True, size=12, color='FFFFFFFF')
    c.fill = fill(C_BLEU_FOND)
    c.alignment = Alignment(horizontal='center', vertical='center')

    wsd.row_dimensions[3].height = 20
    for q in range(1, 29):
        col_letter = get_column_letter(q + 1)
        c = wsd[f'{col_letter}3']
        c.value = f"Q{q}"
        c.font = Font(name='Arial', bold=True, size=10, color='FFFFFFFF')
        c.fill = fill(C_ACCENT)
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = border_full()

    for i, row_data in df.iterrows():
        excel_row = 4 + i
        wsd.row_dimensions[excel_row].height = 16
        for q in range(1, 29):
            col = f'Q{q}'
            val = row_data.get(col, '')
            col_letter = get_column_letter(q + 1)
            bg = 'FFF7F7F7' if i % 2 == 0 else C_BLANC
            c = wsd[f'{col_letter}{excel_row}']
            c.value = val if pd.notna(val) else ''
            c.font = Font(name='Arial', size=9)
            c.fill = fill(bg)
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.border = border_full()

    wb.save(output_path)
    print(f"Rapport généré : {output_path}")
    print(f"Répondants : {n_rep} | Thèmes : {len(THEMES)} | Questions : 28")

# ─────────────────────────────────────────────
if __name__ == '__main__':
    csv_path   = sys.argv[1] if len(sys.argv) > 1 else 'reponses_test.csv'
    out_path   = sys.argv[2] if len(sys.argv) > 2 else 'rapport_rps.xlsx'
    entreprise = sys.argv[3] if len(sys.argv) > 3 else 'Entreprise'
    build_excel(csv_path, out_path, entreprise)
